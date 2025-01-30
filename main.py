import arcpy
import requests
import json
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime
import pytz

# ------------------
# USER PARAMETERS
# ------------------
# Use the feature class from the active ArcGIS Pro project
input_fc = "LakeForestWGS"
lat_field = "Lat"
lon_field = "Lon"

# Setup logging
def log_message(message, level="INFO"):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{timestamp}] {level}: {message}")

# Get project folder
try:
    aprx = arcpy.mp.ArcGISProject("CURRENT")
    project_folder = aprx.homeFolder
    log_message(f"Project folder found: {project_folder}")
except Exception as e:
    log_message(f"Error accessing ArcGIS project: {str(e)}", "ERROR")
    raise

# Excel output settings
excel_filename = "LakeForest_Mapillary_Imagery.xlsx"
output_excel_path = os.path.join(project_folder, excel_filename)
log_message(f"Excel will be saved to: {output_excel_path}")

# Mapillary API settings
mapillary_access_token = "MLY|9441786265842838|7f6f0c2a2d6a89b3aa725bdd2cb34fd0"
bbox_margin = 0.0005
fields = "id,captured_at,geometry,thumb_256_url"

# -----------------------------------------------------------------
# 1. CHECK IF FEATURE CLASS EXISTS
# -----------------------------------------------------------------
log_message(f"Checking for feature class: {input_fc}")
if not arcpy.Exists(input_fc):
    log_message(f"Feature class '{input_fc}' not found!", "ERROR")
    raise RuntimeError(f"Error: The feature class '{input_fc}' was not found in the current ArcGIS Pro project.")
log_message("Feature class found successfully")

# Get the first 50 field names (excluding latitude and longitude)
field_names = [f.name for f in arcpy.ListFields(input_fc) if f.name not in [lat_field, lon_field]]
first_50_fields = field_names[:50]

# -----------------------------------------------------------------
# 2. CHECK IF EXCEL WORKBOOK EXISTS
# -----------------------------------------------------------------
if os.path.exists(output_excel_path):
    wb = load_workbook(output_excel_path)  # Load existing workbook
    ws = wb.active
    log_message("Existing Excel workbook loaded successfully")
else:
    # If it doesn't exist, create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Mapillary Imagery"
    ws.append(["Latitude", "Longitude", "Date", "Image Link", "Image Latitude", "Image Longitude"] + first_50_fields)  # Update header
    log_message("New Excel workbook initialized successfully")

# -----------------------------------------------------------------
# 3. ITERATE OVER FIRST 50 FEATURES
# -----------------------------------------------------------------
count = 0
error_count = 0
success_count = 0

# Get total features but limit to max of 50
total_features = min(int(arcpy.GetCount_management(input_fc)[0]), 50)  # Limit to 50 max
log_message(f"Total features to process (limited): {total_features}")

def safe_str(value):
    if isinstance(value, (tuple, list)):
        return str(value)  # Convert tuple or list to string
    return value

with arcpy.da.SearchCursor(input_fc, [lat_field, lon_field] + first_50_fields) as cursor:  # Include first 50 fields
    for row in cursor:
        count += 1
        latitude = row[0]
        longitude = row[1]
        other_attributes = row[2:]  # Retrieve the additional fields

        log_message(f"Processing feature {count}/{total_features} at coordinates: ({latitude}, {longitude})")

        # Construct bounding box
        left   = longitude - bbox_margin
        bottom = latitude  - bbox_margin
        right  = longitude + bbox_margin
        top    = latitude  + bbox_margin
        bbox_param = f"{left},{bottom},{right},{top}"
        
        # Mapillary search endpoint
        url = "https://graph.mapillary.com/images"
        params = {
            "access_token": mapillary_access_token,
            "bbox": bbox_param,
            "fields": fields,
            "limit": 1
        }
        
        try:
            log_message(f"Making API request for feature {count}")
            resp = requests.get(url, params=params)
            resp.raise_for_status()
            
            # Log API response details
            log_message(f"API Response Status: {resp.status_code}")
            log_message(f"API Response Headers: {dict(resp.headers)}")
            
            data = resp.json()
            images = data.get("data", [])
            
            if not images:
                log_message(f"No images found for coordinates ({latitude}, {longitude})", "WARNING")
                ws.append([latitude, longitude, "No Image Found", "N/A", latitude, longitude] + [safe_str(attr) for attr in other_attributes])  # Append additional fields with coordinates
            else:
                image = images[0]
                captured_at = image.get("captured_at", "Unknown_Date")
                
                if captured_at != "Unknown_Date":
                    # Convert Unix timestamp to datetime object
                    dt = datetime.fromtimestamp(int(captured_at) / 1000, tz=pytz.utc)
                    
                    # Convert to PDT timezone
                    pdt_tz = pytz.timezone('US/Pacific')
                    pdt_dt = dt.astimezone(pdt_tz)
                    
                    # Format the date string
                    captured_at = pdt_dt.strftime('%Y-%m-%d %H:%M:%S %Z%z')
                    
                image_id = image.get("id", "")
                image_link = f"https://www.mapillary.com/app/?focus=photo&pKey={image_id}"
                
                log_message(f"Found image {image_id} captured at {captured_at}")
                ws.append([latitude, longitude, captured_at, image_link, latitude, longitude] + [safe_str(attr) for attr in other_attributes])  # Append additional fields with coordinates
                success_count += 1
                
            # Save periodically
            if count % 10 == 0:
                wb.save(output_excel_path)
                log_message(f"Interim save completed at feature {count}")

            # Break if we have processed 50 features
            if count >= 50:
                break
                
        except requests.exceptions.RequestException as e:
            error_count += 1
            log_message(f"API Error for coordinates ({latitude}, {longitude}): {str(e)}", "ERROR")
            ws.append([latitude, longitude, f"API Error: {e}", "N/A", latitude, longitude] + [safe_str(attr) for attr in other_attributes])  # Append additional fields with coordinates
            
        # Progress update
        if count % 5 == 0:
            progress = (count / total_features) * 100
            log_message(f"Progress: {progress:.1f}% complete")

# -----------------------------------------------------------------
# 4. SAVE THE EXCEL WORKBOOK
# -----------------------------------------------------------------
try:
    wb.save(output_excel_path)
    log_message("Final Excel file save completed successfully")
except Exception as e:
    log_message(f"Error saving Excel file: {str(e)}", "ERROR")
    raise

# Final statistics
log_message("Processing completed!")
log_message(f"Total features processed: {count}")
log_message(f"Successful image retrievals: {success_count}")
log_message(f"Errors encountered: {error_count}")
log_message(f"Excel file saved at: {output_excel_path}")
